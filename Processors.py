import openpyxl as excel

def grade_processor(url, sheet_name):
    file = excel.load_workbook(url)
    sheet = file[sheet_name]

    for row in range(2, sheet.max_row + 1):
        maxobt = sheet.cell(row, 3)
        score = sheet.cell(row, 4)
        percentage = (score.value / maxobt.value) * 100
        pcell = sheet.cell(row, 5)
        pcell.value = percentage
        grade = sheet.cell(row, 6)
        comment = sheet.cell(row, 7)

        if percentage > 75:
            grade.value = 'A'
            comment.value = 'EXCELLENT'
        elif 70 <= percentage < 75:
            grade.value = 'B*'
            comment.value = 'VERY GOOD'
        elif 65 <= percentage < 70:
            grade.value = 'B'
            comment.value = 'GOOD'
        elif 60 <= percentage < 65:
            grade.value = 'c'
            comment.value = 'OMO, YOU LUCKY O'
        else:
            grade.value = 'F'
            comment.value = 'SEE YA NEXT SEM!'

    names = url.split('\\')
    file.save(names[-1])


def salary_processor(url, sheet_name):
    file = excel.load_workbook(url)
    sheet = file[sheet_name]

    for row in range(2, sheet.max_row + 1):
        salary = sheet.cell(row, 2)
        tax = 0.0575 * salary.value
        taxcell = sheet.cell(row, 3)
        taxcell.value = tax
        cooperative = salary.value * 0.3
        C_cell = sheet.cell(row, 4)
        C_cell.value = cooperative
        takehome = salary.value - tax - cooperative
        takehomecell = sheet.cell(row, 5)
        takehomecell.value = takehome

    names = url.split('\\')
    file.save(names[-1])


def sales_processor(url, sheet_name):
    file = excel.load_workbook(url)
    sheet = file[sheet_name]

    for row in range(2, sheet.max_row + 1):
        region = sheet.cell(row, 2)
        quantity = sheet.cell(row, 3)
        ppu = sheet.cell(row, 4)
        shipping = sheet.cell(row, 5)
        TP = sheet.cell(row, 6)
        VAT = sheet.cell(row, 7)
        ATBP = sheet.cell(row, 8)
        total_price = quantity.value * ppu.value
        vat = 0.0575 * total_price

        if region.value == 'Abroad':
            ship = 0.10 * total_price
        else:
            ship = 0

        amount_TBP = vat + ship + total_price
        shipping.value = ship
        TP.value = total_price
        VAT.value = vat
        ATBP.value = amount_TBP

    names = url.split('\\')
    file.save(names[-1])

